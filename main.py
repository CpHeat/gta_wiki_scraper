from pathlib import Path
import csv
import datetime
from itertools import islice
import pandas
import requests
import unicodedata
import re
import bs4
from bs4 import BeautifulSoup
from openpyxl.workbook import Workbook

# Get page content and save it locally
def get_page(url:str, output_file:str):
    print("Scraping url...")
    page = requests.get(url)


    f = open(output_file, "w")
    f.write(str(page.content))
    f.close()


# Normalize filename
def normalize_filename(filename:str):
    filename = unicodedata.normalize('NFKC', filename)
    filename = re.sub(r'[^\w\s-]', '', filename.lower())
    return re.sub(r'[-\s]+', '-', filename).strip('-_')


# Analyze the page and extract the data from the vehicles table
def extract_vehicles_table(vehicles_page:str):
    with open(vehicles_page, "r") as file:
        soup = BeautifulSoup(file, 'html.parser')
        return soup.find_all("table", class_="wikitable")


# Analyze the vehicles table and extract the name and link for every model
def extract_vehicles_list(vehicles_table:str):
    soup = BeautifulSoup(vehicles_table, 'html.parser')
    vehicles = soup.find_all("li")
    vehicles_list = []

    # Some vehicles have no link, detect them and scrape them accordingly
    for vehicle in vehicles:
        if vehicle.find("span") and vehicle.find("span").get("data-uncrawlable-url"):
            vehicle_name = vehicle.find("span").get("title").replace(" (page does not exist)", "")
            vehicle_link = ""
        else:
            vehicle_name = vehicle.find("a").get('title', 'No title attribute')
            vehicle_link = "https://gta.fandom.com" + vehicle.find("a").get('href')

        vehicles_list.append({"name": vehicle_name, "link": vehicle_link})
    print(vehicles_list)
    return vehicles_list


def get_vehicle_image(vehicle_infos:bs4.element.Tag):
    vehicle_image_wrapper = vehicle_infos.find("figure", attrs={'data-source': re.compile(r'^front_image')})
    return vehicle_image_wrapper.find("img").get("src")


def get_vehicle_class(vehicle_infos):
    vehicle_class = vehicle_class_wrapper = vehicle_infos.find("div", attrs={'data-source': 'class'})
    if vehicle_class_wrapper:
        if vehicle_class_wrapper.find("a"):
            vehicle_class = vehicle_class_wrapper.find("a").getText().split(" (")[0]
        else:
            vehicle_class = vehicle_class_wrapper.find("div", class_="pi-font").getText().split(" (")[0]
    return vehicle_class


def get_vehicle_type(vehicle_infos):
    vehicle_type = vehicle_type_wrapper = vehicle_infos.find("div", attrs={'data-source': 'type'})
    if vehicle_type_wrapper:
        vehicle_type = vehicle_type_wrapper.find("div", class_="pi-data-value").getText()
    return vehicle_type


def get_vehicle_body_style(vehicle_infos):
    vehicle_body_style = vehicle_body_style_wrapper = vehicle_infos.find("div", attrs={'data-source': 'body_style'})
    if vehicle_body_style_wrapper:
        vehicle_body_style = vehicle_body_style_wrapper.find("div", class_="pi-data-value").getText()
    return vehicle_body_style


def get_vehicle_capacity(vehicle_infos):
    vehicle_capacity_wrapper = vehicle_infos.find("div", attrs={'data-source': 'capacity'})
    if vehicle_capacity_wrapper:
        return vehicle_capacity_wrapper.find("div", class_="pi-data-value").getText().split(" ")[0]
    else:
        return 0


def get_vehicle_handling(soup:bs4.BeautifulSoup):
    max_velocity_header = soup.find("th", string="Max velocity")
    if max_velocity_header:
        row = max_velocity_header.find_parent("tr")
        if row:
            vehicle_handling_infos_row = row.find_next_sibling("tr")
            vehicle_handling_infos = vehicle_handling_infos_row.find_all("td")
            vehicle_speed = vehicle_handling_infos[0].text.strip("\\n")
            vehicle_drivetrain = vehicle_handling_infos[3].text.strip("\\n")
            return vehicle_speed, vehicle_drivetrain
    else:
        return "", ""


# Get the number of modifications available
def get_modifications(soup:bs4.BeautifulSoup, modifications_list:list):
    modifications = {"total": 0}
    modification_header = soup.find("th", string=lambda text: "Modification" in text if text else False)
    if modification_header:
        modifications_count = 0
        table = modification_header.find_parent("tbody")

        if table:
            for modification in modifications_list:
                modification_cell = table.find("td", string=lambda text: modification in text if text else False, attrs={"rowspan": True})
                if modification_cell:
                    modification_count = int(modification_cell.get("rowspan"))
                    modifications_count += modification_count
                    modifications[modification.strip("\\n")] = modification_count

        modifications["total"] = modifications_count
        return modifications


# Extract the infos for a vehicle
def extract_vehicle_infos(vehicle_page:str, modifications_list:list):
    with open(vehicle_page, "r") as file:
        soup = BeautifulSoup(file, "html.parser")
        vehicle_infos = soup.find(class_ = "pi-theme-gta-with-subtitle")
        vehicle_image = get_vehicle_image(vehicle_infos)
        vehicle_class = get_vehicle_class(vehicle_infos)
        vehicle_type = get_vehicle_type(vehicle_infos)
        vehicle_body_style = get_vehicle_body_style(vehicle_infos)
        vehicle_capacity = get_vehicle_capacity(vehicle_infos)
        vehicle_speed, vehicle_drivetrain = get_vehicle_handling(soup)
        vehicle_modifications = get_modifications(soup, modifications_list)

        return {
            "image": vehicle_image,
            "class": vehicle_class,
            "type": vehicle_type,
            "body-style": vehicle_body_style,
            "speed": vehicle_speed,
            "drivetrain": vehicle_drivetrain,
            "capacity": vehicle_capacity,
            "modifications": vehicle_modifications
        }


# Save infos in a csv file
def load_vehicles_infos(vehicles_list:list, vehicle_csv_output:str):
    with open(vehicle_csv_output, "w", newline='', encoding='utf-8') as output:
        fieldnames = ["name", "link", "image", "class", "type", "body-style", "capacity", "modifications"]
        writer = csv.DictWriter(output, fieldnames = fieldnames, delimiter = ",")
        writer.writeheader()
        for vehicle_data in vehicles_list:
            writer.writerow(vehicle_data)


# Generate a xlsx file with hyperlinks
def generate_xlsx(csv_file:str, output_xlsx:str):
    # load CSV
    df = pandas.read_csv(csv_file)

    # Create Excel file
    wb = Workbook()
    ws = wb.active

    # Add headers
    for col_idx, col_name in enumerate(df.columns, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Insert data with hyperlinks
    for row_idx, row in df.iterrows():
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx + 2, column=col_idx, value=value)

            # Check if value is hyperlink
            if isinstance(value, str) and value.startswith(("http://", "https://")):
                cell.hyperlink = value
                cell.value = "Link"  # Texte raccourci affiché
                cell.style = "Hyperlink"

    # Save Excel file
    wb.save(output_xlsx)
    print(f"Fichier Excel avec hyperliens créé : {output_xlsx}")


def get_cache(cache_file:str, refresh_cache_in_hours:float):
    f = open(cache_file, "r")
    cache_timestamp = float(f.read())
    current_timestamp = datetime.datetime.now().timestamp()
    if cache_timestamp <= (current_timestamp + (refresh_cache_in_hours * 3600)):
        print("Cache outdated, let's go scraping...")
        return True
    else:
        return False


def set_cache(cache_file:str):
    current_timestamp = str(datetime.datetime.now().timestamp())
    f = open(cache_file, "w")
    f.write(current_timestamp)
    f.close()


def main():
    refresh_cache_in_hours:float = 168
    vehicles_page_url = "https://gta.fandom.com/wiki/Vehicles_in_GTA_Online"
    Path("scraped/").mkdir(parents=True, exist_ok=True)
    Path("output/").mkdir(parents=True, exist_ok=True)
    vehicles_page_output = "scraped/gta_vehicles_page.html"
    vehicle_csv_output = "output/vehicles.csv"
    vehicles_xlsx_output = "output/vehicles.xlsx"
    cache_file = "cache.txt"
    refresh_cache = get_cache(cache_file, refresh_cache_in_hours)

    modifications_list = [
        "Armor Plating\\n",
        "Blades\\n",
        "Cam Cover\\n",
        "Dash\\n",
        "Doors\\n",
        "Drift Tuning\\n",
        "Engine Block\\n",
        "Exhaust\\n",
        "Exhausts\\n",
        "Explosives\\n",
        "Fenders\\n",
        "Fog Lights\\n",
        "Front Bumpers\\n",
        "Headlight Covers\\n",
        "Hood Catches\\n",
        "Hoods\\n",
        "Livery\\n",
        "Mirrors\\n",
        "Mudguards\\n",
        "Plateholders\\n",
        "Rear Bumpers\\n",
        "Roll Cages\\n",
        "Roofs\\n",
        "Seats\\n",
        "Skirts\\n",
        "Spikes\\n",
        "Spoilers\\n",
        "Steering Wheels\\n",
        "Strut Braces\\n",
        "Trunks\\n",
        "Vertical Jump\\n",
        "Weapons\\n",
    ]

    # If outdated, get content of the main vehicles page on GTA Wiki
    if refresh_cache:
        get_page(vehicles_page_url, vehicles_page_output)

    # Analyze the page and extract the vehicles list
    vehicles_table = str(extract_vehicles_table(vehicles_page_output))
    vehicles_list = extract_vehicles_list(vehicles_table)

    # Go through the list and get each page. processing_start allow to start at a specific index if needed
    processing_start = 0
    processing_stop = None
    for index, vehicle in enumerate(islice(vehicles_list, processing_start, processing_stop)):
        print(f"Processing {index}: {vehicle}...")

        # If the vehicle has a page get it and extract the infos
        if vehicle["link"]:
            normalized_filename = normalize_filename(vehicle["name"])
            scraped_page_filename = f"scraped/{normalized_filename}.html"
            if refresh_cache:
                get_page(vehicle["link"], scraped_page_filename)

            vehicle_infos = extract_vehicle_infos(scraped_page_filename, modifications_list)
            vehicle["image"] = vehicle_infos["image"]
            vehicle["class"] = vehicle_infos["class"]
            vehicle["type"] = vehicle_infos["type"]
            vehicle["body-style"] = vehicle_infos["body-style"]
            vehicle["capacity"] = vehicle_infos["capacity"]
            vehicle["modifications"] = vehicle_infos["modifications"]
            print(f"Done! {vehicle}")

    print(vehicles_list)
    load_vehicles_infos(vehicles_list, vehicle_csv_output)
    generate_xlsx(vehicle_csv_output, vehicles_xlsx_output)
    set_cache(cache_file)


if __name__ == "__main__":
    main()